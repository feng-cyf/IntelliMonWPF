from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

def auto_adjust_column_width(file_path: str, sheet_name: str = "Sheet1"):
    """
    自适应列宽（openpyxl 版）
    中文、字母、数字混合也能基本对齐
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)   # 列号转 A/B/...
        for cell in col:
            if cell.value:
                length = sum(2 if '\u4e00' <= ch <= '\u9fff' else 1 for ch in str(cell.value))
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)